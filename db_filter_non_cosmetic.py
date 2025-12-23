#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
db_filter_non_cosmetic.py - 过滤非美妆产品记录

功能：
1. 读取翻译后的Excel文件
2. 使用OpenAI API检查F列（产品标题）和G列（翻译）是否为美妆产品
3. 对非美妆产品或空值标记为 "Not Found And No Search"
4. 更新数据库记录
"""

import sqlite3
import openpyxl
import requests
import json
import time
from tqdm import tqdm
import os
import random
from openai import OpenAI
import queue
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed

# 创建OpenAI客户端
client = OpenAI(
    api_key="sk-661023d5822f4b8dbe101d77380f5d1d",
    base_url="https://dashscope.aliyuncs.com/compatible-mode/v1",
)
MODEL_NAME = "qwen3-max"

# 配置
DATABASE_NAME = 'barcode_cache.db'
EXCEL_FILE = '结果old_failures_to_reprocess_translated (1).xlsx'
RESULT_COLUMN = 'H'  # 结果标记列
TITLE_COLUMN = 'F'   # 产品标题列
TRANSLATION_COLUMN = 'G'  # 翻译列
EAN_COLUMN = 'A'     # EAN条码列

# OpenAI API 配置（请替换为实际值）
OPENAI_API_URL = "https://dashscope.aliyuncs.com/compatible-mode/v1"
OPENAI_API_KEY = "sk-661023d5822f4b8dbe101d77380f5d1d"
OPENAI_MODEL = "qwen3-max"

def is_cosmetic_product(title, translation):
    """
    调用OpenAI API判断产品是否为美妆产品
    
    Args:
        title (str): 产品标题
        translation (str): 翻译后的标题
    
    Returns:
        bool: 是否为美妆产品
    """
    # 检查空值
    if not title and not translation:
        return False
    
    # 构建提示词
    prompt = f"""
    请判断以下产品是否为美妆产品（化妆品、护肤品、香水等）。公司主要做美妆产品，所有产品代码都应指向美妆产品。

    产品标题: {title or ''}
    翻译标题: {translation or ''}
    
    请只回答"是"或"否"，不要添加其他内容。
    """
    
    # 调用百炼API
    try:
        # 添加请求延迟，适配百炼API的速率限制
        # 实测速率只有1.27it/s，尝试减少延迟以提高处理速度
        #time.sleep(0.1)  # 每次请求间隔100毫秒（尝试提高到10RPS）
        
        # 性能计时
        start_time = time.time()
        
        max_retries = 5
        base_delay = 1  # 初始延迟1秒
        
        for retry in range(max_retries):
            try:
                completion = client.chat.completions.create(
                    model=MODEL_NAME,
                    messages=[
                        {"role": "system", "content": "你是一个美妆产品分类专家。"},
                        {"role": "user", "content": prompt}
                    ],
                    temperature=0.1
                )
                
                # 解析API响应
                answer = completion.choices[0].message.content.strip().lower()
                return "是" in answer
                
            except Exception as e:
                if "429" in str(e) or "rate limit" in str(e).lower():
                    # 计算指数退避延迟
                    delay = base_delay * (2 ** retry) + random.uniform(0, 1)
                    print(f"收到429错误，等待 {delay:.2f} 秒后重试 ({retry+1}/{max_retries})")
                    time.sleep(delay)
                    continue
                else:
                    raise e
        
        # 如果所有重试都失败
        print("多次重试后仍无法调用百炼API")
        return False
        
    except Exception as e:
        print(f"调用百炼API失败: {e}")
        # API调用失败时，保守处理 - 标记为非美妆产品
        return False

def process_excel_file():
    """处理Excel文件，标记非美妆产品"""
    # 检查文件是否存在
    if not os.path.exists(EXCEL_FILE):
        print(f"错误：Excel文件 {EXCEL_FILE} 不存在")
        return []
    
    # 加载工作簿
    print(f"正在加载Excel文件: {EXCEL_FILE}")
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    
    # 获取列索引
    title_col = openpyxl.utils.column_index_from_string(TITLE_COLUMN)
    trans_col = openpyxl.utils.column_index_from_string(TRANSLATION_COLUMN)
    result_col = openpyxl.utils.column_index_from_string(RESULT_COLUMN)
    ean_col = openpyxl.utils.column_index_from_string(EAN_COLUMN)
    
    # 统计信息
    total_rows = 0
    empty_count = 0
    non_cosmetic_count = 0
    cosmetic_count = 0
    error_count = 0
    
    # 用于存储需要更新的条码
    barcodes_to_update = []
    
    # 遍历所有行（从第2行开始，跳过标题行）
    print("开始处理行数据...")
    
    # 创建线程安全的队列用于存储处理结果
    results_queue = queue.Queue()
    lock = threading.Lock()
    
    # 定义单行处理函数
    def process_single_row(row):
        row_start_time = time.time()
        
        # 获取单元格值
        title = ws.cell(row=row, column=title_col).value
        translation = ws.cell(row=row, column=trans_col).value
        ean = ws.cell(row=row, column=ean_col).value
        
        # 检查空值 - 只需F列（产品标题）为空
        if not title:
            with lock:
                empty_count_local = 1
                barcodes_to_update_local = [(str(ean), "Not Found And No Search")] if ean else []
                
                # 每10行输出一次详细信息
                if (row - 2) % 10 == 0:
                    print(f"t:{title or 'N/A'} tran:{translation or 'N/A'} 判定：空值")
        else:
            try:
                is_cosmetic = is_cosmetic_product(title, translation)
                with lock:
                    if is_cosmetic:
                        cosmetic_count_local = 1
                        empty_count_local = 0
                        barcodes_to_update_local = []
                        
                        # 每10行输出一次详细信息
                        if (row - 2) % 10 == 0:
                            print(f"t:{title or 'N/A'} tran:{translation or 'N/A'} 判定：美妆产品")
                    else:
                        non_cosmetic_count_local = 1
                        empty_count_local = 0
                        barcodes_to_update_local = [(str(ean), "Not Found And No Search")] if ean else []
                        
                        # 每10行输出一次详细信息
                        if (row - 2) % 10 == 0:
                            print(f"t:{title or 'N/A'} tran:{translation or 'N/A'} 判定：非美妆产品")
            except Exception as e:
                with lock:
                    error_count_local = 1
                    empty_count_local = 0
                    barcodes_to_update_local = []
                    
                    # 每10行输出一次详细信息
                    if (row - 2) % 10 == 0:
                        print(f"t:{title or 'N/A'} tran:{translation or 'N/A'} 判定：处理错误 - {str(e)}")
        
        # 将结果放入队列
        results_queue.put((
            row,
            empty_count_local,
            cosmetic_count_local if 'cosmetic_count_local' in locals() else 0,
            non_cosmetic_count_local if 'non_cosmetic_count_local' in locals() else 0,
            error_count_local if 'error_count_local' in locals() else 0,
            barcodes_to_update_local
        ))
        
        # 返回处理时间
        return time.time() - row_start_time
    
    # 使用3个线程处理
    num_threads = 3
    rows = list(range(2, ws.max_row + 1))
    total_rows = len(rows)
    
    # 使用线程池处理
    with ThreadPoolExecutor(max_workers=num_threads) as executor:
        # 提交所有任务
        future_to_row = {executor.submit(process_single_row, row): row for row in rows}
        
        # 使用tqdm显示进度
        for future in tqdm(as_completed(future_to_row), total=total_rows, desc="处理进度"):
            row = future_to_row[future]
            try:
                row_time = future.result()
            except Exception as e:
                print(f"处理行 {row} 时出错: {e}")
    
    # 收集所有结果
    empty_count = 0
    cosmetic_count = 0
    non_cosmetic_count = 0
    error_count = 0
    barcodes_to_update = []
    
    while not results_queue.empty():
        row, ec, cc, ncc, errc, btu = results_queue.get()
        empty_count += ec
        cosmetic_count += cc
        non_cosmetic_count += ncc
        error_count += errc
        barcodes_to_update.extend(btu)
        
        # 更新Excel（在主线程中）
        # 只有空值行和非美妆产品行才标记为 "Not Found And No Search"
        if ec or ncc:  # 空值行或非美妆产品行
            ws.cell(row=row, column=result_col, value="Not Found And No Search")
    
    # 保存Excel文件
    save_start = time.time()
    print("正在保存Excel文件...")
    output_file = EXCEL_FILE.replace('.xlsx', '_filtered.xlsx')
    wb.save(output_file)
    wb.close()
        
    # 打印统计信息
    print(f"Excel保存耗时: {time.time() - save_start:.2f}秒")
    print("\n处理完成！统计信息:")
    print(f"总行数: {total_rows}")
    print(f"空值行数: {empty_count}")
    print(f"美妆产品行数: {cosmetic_count}")
    print(f"非美妆产品行数: {non_cosmetic_count}")
    print(f"错误行数: {error_count}")
    print(f"已保存结果到: {output_file}")
    
    return barcodes_to_update

def update_database(barcodes):
    """更新数据库记录"""
    if not barcodes:
        print("没有需要更新的数据库记录")
        return
    
    conn = sqlite3.connect(DATABASE_NAME)
    cursor = conn.cursor()
    
    updated_count = 0
    
    try:
        for barcode, status in barcodes:
            cursor.execute(
                'UPDATE products SET product_name = ? WHERE barcode = ?',
                (status, barcode)
            )
            if cursor.rowcount > 0:
                updated_count += 1
        
        conn.commit()
        print(f"\n成功更新 {updated_count} 条数据库记录")
    except Exception as e:
        print(f"更新数据库失败: {e}")
        conn.rollback()
    finally:
        conn.close()

def main():
    """主函数"""
    print("=" * 80)
    print("美妆产品过滤工具")
    print("=" * 80)
    print()
    
    # 处理Excel文件
    barcodes_to_update = process_excel_file()
    
    # 更新数据库
    if barcodes_to_update:
        print("\n正在更新数据库...")
        update_database(barcodes_to_update)
        print("\n处理完成！")
    else:
        print("\n没有需要更新的数据库记录")

if __name__ == "__main__":
    main()
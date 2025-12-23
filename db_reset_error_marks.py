#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
db_reset_error_marks.py - 重置错误标记的条码记录

功能：
1. 从 Excel 文件读取需要重置的条码列表
2. 将这些条码的标记从 "Not Found And No Search" 或 "N/A And No Search" 重置为 "Not Found" 或 "N/A"
3. 或者直接删除这些记录，让系统重新检索
"""

import sqlite3
import openpyxl

DATABASE_NAME = 'barcode_cache.db'

def read_barcodes_from_excel(excel_file, sheet_name='条码列表', start_row=2, end_row=None):
    """
    从 Excel 文件读取条码列表
    
    Args:
        excel_file: Excel 文件路径
        sheet_name: 工作表名称，默认为'条码列表'
        start_row: 开始行号（从1开始），默认为2（跳过表头）
        end_row: 结束行号，None表示读取到最后
    
    Returns:
        list: 条码列表
    """
    try:
        wb = openpyxl.load_workbook(excel_file)
        
        # 尝试获取指定的工作表
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            # 如果指定工作表不存在，使用第一个工作表
            ws = wb.active
            print(f"警告：工作表 '{sheet_name}' 不存在，使用第一个工作表: {ws.title}")
        
        barcodes = []
        
        # 确定结束行
        if end_row is None:
            end_row = ws.max_row
        
        # 读取条码（假设条码在第一列）
        for row in range(start_row, end_row + 1):
            barcode = ws.cell(row=row, column=1).value
            if barcode:
                barcodes.append(str(barcode).strip())
        
        wb.close()
        print(f"从 Excel 文件读取到 {len(barcodes)} 个条码")
        return barcodes
        
    except Exception as e:
        print(f"读取 Excel 文件失败: {e}")
        return []

def reset_marks_to_old(barcodes):
    """
    将条码的新标记重置为旧标记
    "Not Found And No Search" -> "Not Found"
    "N/A And No Search" -> "N/A"
    """
    conn = sqlite3.connect(DATABASE_NAME)
    cursor = conn.cursor()
    
    reset_count = 0
    not_found_count = 0
    na_count = 0
    
    try:
        for barcode in barcodes:
            # 查询当前标记
            cursor.execute('SELECT product_name FROM products WHERE barcode = ?', (barcode,))
            result = cursor.fetchone()
            
            if result:
                current_mark = result[0]
                new_mark = None
                
                if current_mark == "Not Found And No Search":
                    new_mark = "Not Found"
                    not_found_count += 1
                elif current_mark == "N/A And No Search":
                    new_mark = "N/A"
                    na_count += 1
                
                if new_mark:
                    cursor.execute(
                        'UPDATE products SET product_name = ? WHERE barcode = ?',
                        (new_mark, barcode)
                    )
                    reset_count += 1
                    print(f"重置条码 {barcode}: {current_mark} -> {new_mark}")
        
        conn.commit()
        print(f"\n✓ 成功重置 {reset_count} 个条码的标记")
        print(f"  - Not Found And No Search -> Not Found: {not_found_count} 个")
        print(f"  - N/A And No Search -> N/A: {na_count} 个")
        
    except Exception as e:
        print(f"错误：重置失败：{e}")
        conn.rollback()
    finally:
        conn.close()

def delete_barcode_records(barcodes):
    """
    直接删除条码记录，让系统重新检索
    """
    conn = sqlite3.connect(DATABASE_NAME)
    cursor = conn.cursor()
    
    deleted_count = 0
    
    try:
        for barcode in barcodes:
            cursor.execute('DELETE FROM products WHERE barcode = ?', (barcode,))
            if cursor.rowcount > 0:
                deleted_count += 1
                print(f"删除条码 {barcode} 的记录")
        
        conn.commit()
        print(f"\n✓ 成功删除 {deleted_count} 个条码记录")
        print("这些条码将在下次运行 backend.py 时重新检索")
        
    except Exception as e:
        print(f"错误：删除失败：{e}")
        conn.rollback()
    finally:
        conn.close()

def main():
    """主函数"""
    print("=" * 80)
    print("错误标记重置工具")
    print("=" * 80)
    print()
    
    # 获取 Excel 文件路径
    excel_file = input("请输入 Excel 文件路径（默认: old_failures_to_reprocess.xlsx）: ").strip()
    if not excel_file:
        excel_file = "old_failures_to_reprocess.xlsx"
    
    # 获取行号范围
    start_row_input = input("请输入开始行号（默认: 2）: ").strip()
    start_row = int(start_row_input) if start_row_input else 2
    
    end_row_input = input("请输入结束行号（默认: 153，输入0表示读取到最后）: ").strip()
    if end_row_input and int(end_row_input) > 0:
        end_row = int(end_row_input)
    else:
        end_row = None
    
    # 读取条码
    print(f"\n正在从 {excel_file} 读取条码...")
    barcodes = read_barcodes_from_excel(excel_file, start_row=start_row, end_row=end_row)
    
    if not barcodes:
        print("未读取到任何条码，程序退出")
        return
    
    print(f"\n读取到 {len(barcodes)} 个条码")
    print(f"前5个条码: {barcodes[:5]}")
    
    # 选择操作
    print("\n请选择操作:")
    print("1. 重置标记（Not Found And No Search -> Not Found, N/A And No Search -> N/A）")
    print("2. 直接删除记录（让系统重新检索）")
    print("0. 退出")
    
    choice = input("\n请输入选项 (0-2): ").strip()
    
    if choice == '1':
        confirm = input(f"\n确认重置 {len(barcodes)} 个条码的标记？(yes/no): ").strip().lower()
        if confirm == 'yes':
            reset_marks_to_old(barcodes)
            print("\n提示：现在可以运行 backend.py 重新处理这些条码")
        else:
            print("操作已取消")
    elif choice == '2':
        confirm = input(f"\n确认删除 {len(barcodes)} 个条码的记录？(yes/no): ").strip().lower()
        if confirm == 'yes':
            delete_barcode_records(barcodes)
            print("\n提示：现在可以运行 backend.py 重新处理这些条码")
        else:
            print("操作已取消")
    else:
        print("已退出")

if __name__ == "__main__":
    main()
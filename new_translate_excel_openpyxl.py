import openpyxl
from openai import OpenAI
from tqdm import tqdm
import os
from concurrent.futures import ThreadPoolExecutor, as_completed
import time
import threading
from threading import Lock
from queue import Queue

# 匹配香水时代名称
from difflib import SequenceMatcher
from rapidfuzz import fuzz
import pandas as pd
import json
class PerfumeMatcher:
    def __init__(self, term_file):
        self.df = pd.read_excel(term_file)
        self._preprocess_dataframe()

    def _preprocess_dataframe(self):
        """一次性预处理所有香水数据"""
        self.df['full_name'] = self.df.iloc[:, 0].astype(str).str.replace('[,，]', ' ', regex=True).str.strip()

        # 处理末尾4位数字（如年份）
        self.df['full_name'] = self.df['full_name'].apply(
            lambda x: x[:-4].strip() if len(x) >= 4 and x[-4:].isdigit() else x
        )

        self.df['en_name'] = self.df['full_name'].apply(
            lambda x: ' '.join([p for p in x.split() if p.isascii()]).lower()
        )
        self.df['cn_name'] = self.df['full_name'].apply(
            lambda x: x.split()[0] if x.split() else ''
        )

    def preprocess(self, text):
        """清洗用户输入"""
        return text.lower().strip()

    def get_all_matches(self, text, threshold=0.3):
        """返回相似度大于阈值的前5个结果"""
        start_time = time.time()
        query = self.preprocess(text)

        def match_row(row):
            score = fuzz.ratio(query, row['en_name']) / 100.0
            if score >= threshold:
                return {
                    '中文名': row['cn_name'],
                    '原始名称': row['full_name'],
                    '匹配分数': score,
                    '输入查询': text,
                    '响应时间(ms)': round((time.time() - start_time) * 1000, 2)
                }
            return None

        matches = filter(None, self.df.apply(match_row, axis=1).tolist())
        matches = sorted(matches, key=lambda x: x['匹配分数'], reverse=True)
        return matches[:5] if matches else None

    def get_top_names(self, text, threshold=0.3):
        """返回匹配度最高的名称列表"""
        matches = self.get_all_matches(text, threshold)
        return [m['原始名称'] for m in matches] if matches else None

    def get_CHINESE_NAME(self, ean):
        """根据ean条码从中文品名汇总表.xlsx获取中文名称"""
        try:
            # 这里需要实现从中文品名汇总表.xlsx中根据ean查找中文名称的逻辑
            # 示例代码 - 实际实现需要根据具体文件格式调整
            #print(f"正在根据EAN {ean} 查找中文品名...")
            if not hasattr(self, 'ean_df'):#
                self.ean_df = pd.read_excel('中文品名汇总表.xlsx')
                """
                #打印表单行数
                print(f"中文品名汇总表.xlsx 行数: {len(self.ean_df)}")
                print(f"中文品名汇总表.xlsx 列数: {len(self.ean_df.columns)}")
                #打印表单前5行
                print(f"中文品名汇总表.xlsx 前5行数据:")
                print(self.ean_df.head())
                #打印表单14016行和其之后5行
                print(f"中文品名汇总表.xlsx 第14016行及之后5行数据:")
                print(self.ean_df.iloc[14016:14021])
                """

            """
            # 添加数据类型调试
            print(f"EAN数据类型: {type(ean)}, 值: {ean}")
            print(f"条码列数据类型: {self.ean_df['条码'].dtype}")
            """
            # 确保ean是字符串类型进行比较
            ean_str = str(ean).strip()
            #print(f"转换后的EAN: '{ean_str}'")
            
            # 尝试精确匹配
            result = self.ean_df[self.ean_df['条码'].astype(str).str.strip() == ean_str]
            
            if not result.empty:
                print(f"找到EAN {ean} 对应的中文品名: {result.iloc[0]['中文品名']}")
                return result.iloc[0]['中文品名']
            
            # 如果精确匹配失败，尝试打印一些调试信息
            #print(f"精确匹配失败，尝试查看前几个条码:")
            #print(self.ean_df['条码'].astype(str).str.strip().head(10).tolist())
            
            return None
        except Exception as e:
            print(f"EAN查找失败: {str(e)}")
            return None

# 启动读取香水名称数据表
matcher = PerfumeMatcher('all(1).xlsx')

MAX_WORKERS = 15

class APICounter:
    """API调用计数器（线程安全）"""
    def __init__(self):
        self.count = 0
        self._lock = Lock()
    
    def increment(self):
        with self._lock:
            self.count += 1
    
    def reset(self):
        with self._lock:
            count = self.count
            self.count = 0
            return count

api_counter = APICounter()

client = OpenAI(
    api_key=json.load(open("config.json"))["api_key"],#阿里云
    base_url="https://dashscope.aliyuncs.com/compatible-mode/v1",#阿里云
    #base_url="https://api.deepseek.com",
    #api_key="sk-f8fed3abc342479685d308cc8cfc51a6",
)
prompt0 = "任务背景：将提供给你一段美妆产品描述和参考内容（参考内容可能与美妆产品描述不太相关，如果不相关请当做不存在。如果参考内容可能与美妆产品描述相关，请参考参考内容进行处理）。\n任务目标：将美妆产品描述进行整理，格式要求：【品牌名（如有）】 【产品名（如有）】 【规格/毫升数（如有）】其中产品名和规格翻译成中文，要求贴合美妆产品的实际名称（信达雅）{产品名必须翻译为中文}【如果原来的顺序错乱，请调换】【例如：【感官之水】【淡香水】【50毫升】】，不要输出多余的东西"

def api_counter_thread():
    last_time = time.time()
    while True:
        time.sleep(60)
        current_time = time.time()
        count = api_counter.reset()
        elapsed_time = current_time - last_time
        requests_per_second = count / elapsed_time if elapsed_time > 0 else 0
        print(f"[统计] 最近1分钟API请求次数: {count} (平均 {requests_per_second:.2f} 次/秒)")
        last_time = current_time

def translate_single(text, ean=None, pinpai=None):
    """新版翻译函数，支持ean条码参数"""
        # 优先使用ean条码匹配
    if ean:
        print(f"[INFO] 使用EAN条码匹配: {ean}")
        chinese_name = matcher.get_CHINESE_NAME(ean)
        if chinese_name:
            return chinese_name

    if not text or str(text).strip().lower() == 'nan':
        return "NAN"
    

    
    # 无ean或匹配失败时使用香水名称匹配
    name_results = matcher.get_top_names(str(text))
    
    if name_results:
        cankaoneirong = "参考内容：" + ", ".join(name_results)+"\n如果参考内容中的英文与待翻译内容没有任何相似，请不要参考参考内容，品牌名保留英文"
    else:
        cankaoneirong = "参考内容：无"
    
    daifanyineirong = "待翻译内容：" + str(text)
    
    completion = client.chat.completions.create(
        model="deepseek-v3",
        messages=[
            {"role": "system", "content": prompt0},
            {"role": "user", "content": cankaoneirong},
            {"role": "user", "content": daifanyineirong},
        ],
        temperature=0.1,
    )
    api_counter.increment()
    #打印参考内容
    print(f"参考内容：")
    print(name_results)
    print(f"待翻译内容：")
    print(text)
    #print(completion.choices[0].message.content)
    return completion.choices[0].message.content

def process_single(ws, row_idx, src_col_idx, dst_col_idx, barcode_col_idx=None, show_log=True):
    try:
        src_cell = ws.cell(row=row_idx, column=src_col_idx+1)
        original = src_cell.value
        

            
        # 获取ean条码（如果提供了barcode列）
        ean = None
        if barcode_col_idx is not None:
            barcode_cell = ws.cell(row=row_idx, column=barcode_col_idx+1)
            ean = barcode_cell.value if barcode_cell.value else None
            
        try:
            result = translate_single(original, ean)
            ws.cell(row=row_idx, column=dst_col_idx+1, value=result)
        except Exception as e:
            try:
                print(f"第{row_idx}行第一次翻译失败，正在重试...")
                result = translate_single(original, ean)
                ws.cell(row=row_idx, column=dst_col_idx+1, value=result)
            except Exception as e:
                ws.cell(row=row_idx, column=dst_col_idx+1, value="翻译失败")
                if show_log:
                    print(f"第{row_idx}行翻译失败: {str(e)}")
                return False
        
        if show_log:
            print(f"原文: {original}\n翻译: {result}\n{'='*50}")
        return True
    except Exception as e:
        ws.cell(row=row_idx, column=dst_col_idx+1, value="翻译失败")
        if show_log:
            print(f"第{row_idx}行处理失败: {str(e)}")
        return False

def col_name_to_number(col_name):
    result = 0
    for char in col_name.upper():
        result = result * 26 + (ord(char) - ord('A') + 1)
    return result - 1

def col_number_to_name(col_number):
    result = ""
    col_number += 1
    while col_number:
        col_number, remainder = divmod(col_number - 1, 26)
        result = chr(ord('A') + remainder) + result
    return result

def get_excel_row_count(input_path):
    try:
        wb = openpyxl.load_workbook(input_path, read_only=True)
        ws = wb.active
        return ws.max_row
    except Exception as e:
        print(f"读取文件失败: {str(e)}")
        return -1

def translate_excel(input_path, output_path=None, start_row=0, end_row=None,
                  src_col=3, dst_col=11, barcode_col=None, show_log=True):
    #input_path
    """主翻译函数，支持barcode_col参数"""
    if isinstance(src_col, str):
        src_col_idx = col_name_to_number(src_col)
    else:
        src_col_idx = src_col
        
    if isinstance(dst_col, str):
        dst_col_idx = col_name_to_number(dst_col)
    else:
        dst_col_idx = dst_col
        
    # 处理barcode_col参数
    barcode_col_idx = None
    if barcode_col is not None:
        if isinstance(barcode_col, str):
            barcode_col_idx = col_name_to_number(barcode_col)
        else:
            barcode_col_idx = barcode_col

    try:
        wb = openpyxl.load_workbook(input_path)
        ws = wb.active
        
        if output_path is None:
            output_path = input_path.replace('.xlsx', '_translated.xlsx')
            
        total_rows = ws.max_row
        if end_row is None or end_row > total_rows:
            end_row = total_rows
            
        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
            futures = []
            for row_idx in range(start_row, end_row + 1):
                futures.append(executor.submit(
                    process_single, ws, row_idx, src_col_idx, dst_col_idx, barcode_col_idx, show_log
                ))
            
            for future in tqdm(as_completed(futures), total=end_row-start_row+1, desc="翻译进度"):
                future.result()
                
        wb.save(output_path)
        print(f"翻译完成，结果已保存到: {output_path}")
        return {
            "row_count": end_row - start_row + 1,
            "output_path": output_path
        }
        
    except Exception as e:
        print(f"处理失败: {str(e)}")
        return {
            "row_count": 0,
            "output_path": None,
            "error": str(e)
        }

if __name__ == "__main__":
    
    """
    answer=translate_single("LANCOME SET GENIFIQUE FACE 100 & LIGHT P")
    #打印翻译结果
    print("翻译结果：")
    print(f"{answer}") """

    #测试条码匹配
    answer2=matcher.get_CHINESE_NAME("3147758029383")
    answer3=matcher.get_CHINESE_NAME("3346130021148")
    print("条码匹配结果：")
    print(f"{answer2}")
    print(f"{answer3}")

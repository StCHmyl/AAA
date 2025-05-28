

import os
from PIL import Image as PILImage
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException, WebDriverException
from selenium_stealth import stealth
import openpyxl
from openpyxl.drawing.image import Image
# 导入 openpyxl.utils 模块，用于列字母和索引转换
from openpyxl.utils import column_index_from_string
from barcode_excel_db import init_database, get_product_from_db, insert_product_to_db

# --- Excel 操作 ---
# 修改 read_barcodes_from_excel 函数以返回 (barcode, row_index) 对
def read_barcodes_from_excel_with_row(filepath, barcode_column_letter, start_row, end_row):
    """从Excel文件读取指定范围内的条形码，并返回 (barcode, row_index) 对列表。"""
    barcodes_with_row = []
    print(f"读取Excel：开始从文件 {filepath} 的 {barcode_column_letter} 列，第 {start_row} 到 {end_row} 行读取条码。") # 增加日志
    try:
        workbook = openpyxl.load_workbook(filepath)
        sheet = workbook.active
        
        # 将列字母转换为列索引 (A->1, B->2, ...)
        barcode_col_index = column_index_from_string(barcode_column_letter)

        for row_index in range(start_row, end_row + 1):
            cell_value = sheet.cell(row=row_index, column=barcode_col_index).value
            if cell_value is not None and str(cell_value).strip() != "":
                barcodes_with_row.append((str(cell_value).strip(), row_index))
                print(f"读取Excel：读取到条码 {cell_value} 在第 {row_index} 行。") # 增加日志
            else:
                print(f"读取Excel：跳过第 {row_index} 行的空条码。") # 增加日志

    except FileNotFoundError:
        print(f"错误：读取Excel文件失败，文件未找到 - {filepath}") # 增加日志
        return None
    except Exception as e:
        print(f"错误：读取Excel文件时发生未知错误：{e}") # 增加日志
        return None
        
    print(f"读取Excel：成功读取 {len(barcodes_with_row)} 个有效条码及对应行号。") # 增加日志
    return barcodes_with_row

# 修改 update_excel_with_results 函数以接收包含行号的结果列表
def update_excel_with_results_with_row(filepath, results_with_row, barcode_column_letter, product_name_column_letter, image_column_letter):
    """
    根据爬取结果更新Excel文件，写入产品名称并嵌入图片。
    接收包含行号的结果列表。
    """
    print(f"更新Excel：开始更新文件 {filepath}") # 增加日志
    try:
        workbook = openpyxl.load_workbook(filepath)
        sheet = workbook.active
        print(f"更新Excel：成功加载工作簿和活动工作表。") # 增加日志

        # 将列字母转换为列索引
        # 使用导入的 column_index_from_string 函数
        barcode_col_index = column_index_from_string(barcode_column_letter)
        product_name_col_index = column_index_from_string(product_name_column_letter)
        image_col_index = column_index_from_string(image_column_letter)
        print(f"更新Excel：条码列索引: {barcode_col_index}, 产品名称列索引: {product_name_col_index}, 图片列索引: {image_col_index}") # 增加日志

        # 设置图片列的宽度 (在循环外部设置一次)
        sheet.column_dimensions[chr(ord('A') + image_col_index - 1)].width = 12 # 设置图片列的宽度

        # 遍历包含行号的结果列表并更新Excel
        for barcode, product_name, image_filepath, row_index in results_with_row:
            print(f"更新Excel：正在处理条码 {barcode}，对应Excel行号 {row_index}。") # 增加日志
            
            # 设置当前行的行高 (在循环内部为每行设置)
            sheet.row_dimensions[row_index].height = 40 # 设置固定行高

            # 写入产品名称
            sheet.cell(row=row_index, column=product_name_col_index).value = product_name
            print(f"更新Excel：写入产品名称 '{product_name}' 到行 {row_index}，列 {product_name_column_letter}。") # 增加日志

            # 嵌入图片
            if image_filepath and os.path.exists(image_filepath):
                try:
                    # 直接使用Pillow Image对象创建openpyxl Image
                    # 导入 PILImage
                    from PIL import Image as PILImage
                    pil_img = PILImage.open(image_filepath)
                    # 将图片转换为RGB模式(避免alpha通道问题)
                    if pil_img.mode != 'RGB':
                        pil_img = pil_img.convert('RGB')

                    # 获取原图尺寸用于计算比例
                    orig_width, orig_height = pil_img.size

                    # 创建openpyxl图片对象
                    # openpyxl Image可以直接从图片文件路径创建
                    img = Image(image_filepath)

                    # 计算缩放比例,以保持纵横比
                    # 使用 image_col_index 代替硬编码的 15
                    # 需要将列宽和行高转换为像素单位进行计算
                    # openpyxl 的 column_dimensions.width 是以字符宽度为单位的，需要转换
                    # 1个字符宽度大约等于 7 像素 (这是一个经验值，可能因字体和DPI而异)
                    # openpyxl 的 row_dimensions.height 是以磅 (points) 为单位的，需要转换
                    # 1磅 = 1/72 英寸，1英寸 = 96 像素 (标准 DPI)
                    # 1磅 ≈ 96/72 = 1.33 像素
                    # 假设 1字符宽度 ≈ 7像素，1磅 ≈ 1.33像素
                    # 注意：这里假设列宽和行高已经在函数外部设置好
                    col_width_pixels = sheet.column_dimensions[chr(ord('A') + image_col_index - 1)].width * 7
                    row_height_pixels = sheet.row_dimensions[row_index].height * 1.33 # 使用 row_index

                    # 避免除以零
                    if orig_width == 0 or orig_height == 0:
                         print(f"警告：图片 {image_filepath} 尺寸为零，无法嵌入。")
                         sheet.cell(row=row_index, column=image_col_index).value = "Invalid Image Size"
                    else:
                        scale_factor = min(
                            col_width_pixels / orig_width,
                            row_height_pixels / orig_height
                        )

                        # 设置显示尺寸
                        img.width = int(orig_width * scale_factor)
                        img.height = int(orig_height * scale_factor)

                        # 设置图片位置到单元格
                        cell = sheet.cell(row=row_index, column=image_col_index) # 使用 sheet 和 row_index
                        img.anchor = cell.coordinate

                        # 添加图片到工作表
                        sheet.add_image(img) # 使用 sheet
                        print(f"更新Excel：成功嵌入图片 {os.path.basename(image_filepath)} 到行 {row_index}，列 {image_column_letter}。") # 增加日志
                except Exception as e:
                    print(f"错误：嵌入图片 {image_filepath} 到行 {row_index} 失败: {e}") # 增加日志
            elif product_name == "Not Found":
                 # 如果条码未找到，可以在图片列标记一下
                 sheet.cell(row=row_index, column=image_col_index).value = "Image Not Found"
                 print(f"更新Excel：条码 {barcode} 未找到，在图片列标记 'Image Not Found'。") # 增加日志
            else:
                 # 如果爬取成功但没有图片URL或下载失败
                 sheet.cell(row=row_index, column=image_col_index).value = "Image Download Failed"
                 print(f"更新Excel：条码 {barcode} 图片下载失败，在图片列标记 'Image Download Failed'。") # 增加日志


        # 保存更新后的Excel文件
        try:
            workbook.save(filepath)
            print(f"更新Excel：成功保存更新后的Excel文件: {filepath}") # 增加日志
        except Exception as e:
                print(f"错误：保存Excel文件时发生错误：{e}") # 增加日志

    except Exception as e:
        print(f"错误：更新Excel文件时发生未知错误：{e}") # 增加日志

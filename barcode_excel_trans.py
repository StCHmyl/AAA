import openpyxl
from openai import OpenAI
from tqdm import tqdm
import os
from concurrent.futures import ThreadPoolExecutor, as_completed
import time
import threading
from threading import Lock
from queue import Queue
import json
MAX_WORKERS = 30

class APICounter:
    """API调用计数器（线程安全）
    
    属性：
    count : int - 累计API调用次数
    _lock : Lock - 保证线程安全的锁
    
    方法：
    increment() : 增加计数器
    reset() : 重置并返回当前计数值
    """
    def __init__(self):
        """初始化计数器"""
        self.count = 0  # 从0开始计数
        self._lock = Lock()  # 创建线程锁
    
    def increment(self):
        """原子化增加计数器（线程安全）"""
        with self._lock:  # 自动获取和释放锁
            self.count += 1
    
    def reset(self):
        """原子化重置计数器并返回当前值（线程安全）"""
        with self._lock:
            count = self.count
            self.count = 0  # 重置为0
            return count  # 返回重置前的值

api_counter = APICounter()

client = OpenAI(
    api_key=json.load(open("config.json"))["api_key"],
    base_url="https://dashscope.aliyuncs.com/compatible-mode/v1",
)
prompt = "将美妆产品描述翻译成中文，要求贴合美妆产品的实际名称（信达雅）{产品名必须翻译为中文}，格式要求：【品牌名（如有）】 【产品名（如有）】 【规格/毫升数（如有）】【如果原来的顺序错乱，请调换】【例如：【感官之水】【淡香水】【50毫升】】，不要输出多余的东西"

def api_counter_thread():
    """每分钟统计API调用次数的后台线程
    
    工作流程：
    1. 记录初始时间
    2. 每分钟循环执行：
        a. 重置计数器并获取计数值
        b. 计算请求速率
        c. 打印统计信息
    3. 使用daemon模式运行，主程序退出时自动结束
    """
    last_time = time.time()  # 记录初始时间戳
    while True:
        time.sleep(60)  # 精确等待60秒
        current_time = time.time()
        count = api_counter.reset()  # 获取并重置计数器
        elapsed_time = current_time - last_time
        # 计算每秒请求数（处理除零情况）
        requests_per_second = count / elapsed_time if elapsed_time > 0 else 0
        # 格式化输出统计信息
        print(f"[统计] 最近1分钟API请求次数: {count} (平均 {requests_per_second:.2f} 次/秒)")
        last_time = current_time  # 更新时间戳

def translate_single(text):
    if not text or str(text).strip().lower() == 'nan':
        return "NAN"
        
    completion = client.chat.completions.create(
        model="deepseek-v3",
        messages=[
            {"role": "system", "content": prompt},
            {"role": "user", "content": str(text)}
        ],
        temperature=0.3,
    )
    api_counter.increment()
    return completion.choices[0].message.content

def process_single(ws, row_idx, src_col_idx, dst_col_idx, show_log=True):
    try:
        src_cell = ws.cell(row=row_idx, column=src_col_idx+1)
        original = src_cell.value
        
        if not original or str(original).strip() == '':
            ws.cell(row=row_idx, column=dst_col_idx+1, value="NAN")
            return True
            
        try:
            result = translate_single(original)
            ws.cell(row=row_idx, column=dst_col_idx+1, value=result)
        except Exception as e:
            try:
                print(f"第{row_idx}行第一次翻译失败，正在重试...")
                result = translate_single(original)
                ws.cell(row=row_idx, column=dst_col_idx+1, value=result)
            except Exception as e:
                ws.cell(row=row_idx, column=dst_col_idx+1, value="翻译失败")
                if show_log:
                    print(f"第{row_idx}行翻译失败: {str(e)}")
                return False
        
        if show_log:
            print(f"原文: {original}\n翻译: {result}\n{'='*50}")
        return True
    except Exception as e:
        ws.cell(row=row_idx, column=dst_col_idx+1, value="翻译失败")
        if show_log:
            print(f"第{row_idx}行处理失败: {str(e)}")
        return False

def col_name_to_number(col_name):
    result = 0
    for char in col_name.upper():
        result = result * 26 + (ord(char) - ord('A') + 1)
    return result - 1

def col_number_to_name(col_number):
    result = ""
    col_number += 1
    while col_number:
        col_number, remainder = divmod(col_number - 1, 26)
        result = chr(ord('A') + remainder) + result
    return result

def get_excel_row_count(input_path):
    try:
        wb = openpyxl.load_workbook(input_path, read_only=True)
        ws = wb.active
        return ws.max_row
    except Exception as e:
        print(f"读取文件失败: {str(e)}")
        return -1

def translate_excel(input_path, output_path=None, start_row=0, end_row=None,
                  src_col=3, dst_col=11, show_log=True):
    """
    主翻译函数 - 使用openpyxl处理Excel文件
    
    参数说明：
    input_path: 输入文件路径
    output_path: 输出文件路径（默认在原文件名后加_translated）
    start_row: 起始行号（从0开始计数，对应Excel第2行）
    end_row: 结束行号（包含该行）
    src_col: 源列（可以是字母或数字，从0开始）
    dst_col: 目标列（同上）
    show_log: 是否显示详细日志
    
    注意事项：
    1. openpyxl行号从1开始，因此代码中有+1调整
    2. start_row参数从0开始以保持接口统一
    3. 使用多线程处理提高效率（MAX_WORKERS控制并发数）
    """
    
    # 列名/列号转换处理（支持字母和数字两种输入格式）
    if isinstance(src_col, str):
        src_col_idx = col_name_to_number(src_col)  # 将字母列名转为数字索引
    else:
        src_col_idx = src_col  # 直接使用数字索引
        
    if isinstance(dst_col, str):
        dst_col_idx = col_name_to_number(dst_col)
    else:
        dst_col_idx = dst_col

    try:
        wb = openpyxl.load_workbook(input_path)
        ws = wb.active
        
        if output_path is None:
            output_path = input_path.replace('.xlsx', '_translated.xlsx')
            
        total_rows = ws.max_row
        if end_row is None or end_row > total_rows:
            end_row = total_rows
            
        #counter_thread = threading.Thread(target=api_counter_thread, daemon=True)统计每分钟的API调用次数
        # 启动API调用计数器线程
        #counter_thread.start()
        
        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
            futures = []
        
            for row_idx in range(start_row , end_row ):  # openpyxl行号从1开始
                futures.append(executor.submit(
                    process_single, ws, row_idx, src_col_idx, dst_col_idx, show_log
                ))
            
            for future in tqdm(as_completed(futures), total=end_row-start_row+1, desc="翻译进度"):
                future.result()
                
        wb.save(output_path)
        print(f"翻译完成，结果已保存到: {output_path}")
        return {
            "row_count": end_row - start_row + 1,
            "output_path": output_path
        }
        
    except Exception as e:
        print(f"处理失败: {str(e)}")
        return {
            "row_count": 0,
            "output_path": None,
            "error": str(e)
        }

if __name__ == "__main__":
    input_file = "Copia di GIACENZA 1804.XLSX"
    translate_excel(
        input_file,
        start_row=0,
        end_row=498,
        src_col='E',
        dst_col='G',
        show_log=False
    )

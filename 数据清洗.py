from openpyxl import load_workbook

def clean_duplicates(input_file, output_file):
    # 加载工作簿
    wb = load_workbook(input_file)
    ws = wb.active
    
    # 存储已出现的A列值和对应行号
    seen_values = {}
    rows_to_keep = []
    deleted_rows = []
    
    # 保留标题行(第一行)
    rows_to_keep.append(1)
    
    # 从第二行开始检查A列
    for row in range(2, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=1).value
        
        # 如果值未出现过，则保留该行
        if cell_value not in seen_values:
            seen_values[cell_value] = row
            rows_to_keep.append(row)
        else:
            # 记录将被删除的行内容
            row_data = [ws.cell(row=row, column=col).value for col in range(1, ws.max_column + 1)]
            deleted_rows.append((row, row_data))
    
    # 打印将被删除的行
    if deleted_rows:
        print("\n将被删除的重复行内容:")
        print("-" * 80)
        print(f"{'行号':<8}{'内容':<72}")
        print("-" * 80)
        for row_num, row_data in deleted_rows:
            print(f"{row_num:<8}{str(row_data):<72}")
        print("-" * 80)
        print(f"共发现 {len(deleted_rows)} 行重复数据将被删除\n")
    
    # 创建新工作簿
    new_wb = load_workbook(input_file)
    new_ws = new_wb.active
    
    # 删除不需要的行(从后往前删除避免索引问题)
    for row in sorted(set(range(1, ws.max_row + 1)) - set(rows_to_keep), reverse=True):
        new_ws.delete_rows(row)
    
    # 保存到新文件
    new_wb.save(output_file)
    print(f"数据清洗完成，结果已保存到: {output_file}")

# 使用示例
input_file = "中文品名汇总表.xlsx"
output_file = "中文品名汇总表_cleaned.xlsx"
clean_duplicates(input_file, output_file)

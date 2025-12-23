#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
db_list_old_failures.py - 列出需要重新检索的旧失败记录

功能：
1. 查找标记为 "N/A" 或 "Not Found" 的旧记录
2. 导出清单到 Excel 文件（可直接用于系统处理）
3. 提供批量重置选项，将这些记录标记为待重新检索
"""

import sqlite3
import json
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

DATABASE_NAME = 'barcode_cache.db'

def get_old_failure_records():
    """获取所有旧的失败记录（N/A 和 Not Found）"""
    conn = sqlite3.connect(DATABASE_NAME)
    cursor = conn.cursor()
    
    # 查找 product_name 为 "N/A" 或 "Not Found" 的记录
    cursor.execute('''
        SELECT barcode, product_name, image_url, image_filepath
        FROM products 
        WHERE product_name IN ('N/A', 'Not Found')
        ORDER BY product_name, barcode
    ''')
    
    records = cursor.fetchall()
    conn.close()
    
    return records

def categorize_records(records):
    """将记录按类型分类"""
    na_records = []
    not_found_records = []
    
    for barcode, product_name, image_url, image_filepath in records:
        record = {
            'barcode': barcode,
            'product_name': product_name,
            'image_url': image_url,
            'image_filepath': image_filepath
        }
        
        if product_name == 'N/A':
            na_records.append(record)
        elif product_name == 'Not Found':
            not_found_records.append(record)
    
    return na_records, not_found_records

def print_summary(na_records, not_found_records):
    """打印统计摘要"""
    total = len(na_records) + len(not_found_records)
    
    print("=" * 80)
    print("旧失败记录统计")
    print("=" * 80)
    print(f"总计: {total} 条记录需要重新检索")
    print(f"  - N/A 记录: {len(na_records)} 条")
    print(f"  - Not Found 记录: {len(not_found_records)} 条")
    print("=" * 80)
    print()

def export_to_excel(na_records, not_found_records, filename='old_failures_to_reprocess.xlsx'):
    """导出清单到 Excel 文件（可直接用于 backend.py 处理）"""
    wb = openpyxl.Workbook()
    
    # 删除默认工作表
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    # 创建汇总工作表
    ws_summary = wb.create_sheet(title='汇总')
    
    # 设置标题样式
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    
    # 写入汇总信息
    ws_summary['A1'] = '旧失败记录汇总'
    ws_summary['A1'].font = Font(size=14, bold=True)
    ws_summary['A3'] = '导出时间:'
    ws_summary['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    ws_summary['A4'] = '总记录数:'
    ws_summary['B4'] = len(na_records) + len(not_found_records)
    ws_summary['A5'] = 'N/A 记录:'
    ws_summary['B5'] = len(na_records)
    ws_summary['A6'] = 'Not Found 记录:'
    ws_summary['B6'] = len(not_found_records)
    
    # 创建 N/A 记录工作表
    if na_records:
        ws_na = wb.create_sheet(title='NA记录')
        
        # 写入表头
        headers = ['条码', '产品名称', '图片URL', '图片路径']
        for col_num, header in enumerate(headers, 1):
            cell = ws_na.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # 写入数据
        for row_num, record in enumerate(na_records, 2):
            ws_na.cell(row=row_num, column=1, value=record['barcode'])
            ws_na.cell(row=row_num, column=2, value=record['product_name'])
            ws_na.cell(row=row_num, column=3, value=record['image_url'])
            ws_na.cell(row=row_num, column=4, value=record['image_filepath'])
        
        # 调整列宽
        ws_na.column_dimensions['A'].width = 15
        ws_na.column_dimensions['B'].width = 15
        ws_na.column_dimensions['C'].width = 40
        ws_na.column_dimensions['D'].width = 40
    
    # 创建 Not Found 记录工作表
    if not_found_records:
        ws_nf = wb.create_sheet(title='NotFound记录')
        
        # 写入表头
        headers = ['条码', '产品名称', '图片URL', '图片路径']
        for col_num, header in enumerate(headers, 1):
            cell = ws_nf.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # 写入数据
        for row_num, record in enumerate(not_found_records, 2):
            ws_nf.cell(row=row_num, column=1, value=record['barcode'])
            ws_nf.cell(row=row_num, column=2, value=record['product_name'])
            ws_nf.cell(row=row_num, column=3, value=record['image_url'])
            ws_nf.cell(row=row_num, column=4, value=record['image_filepath'])
        
        # 调整列宽
        ws_nf.column_dimensions['A'].width = 15
        ws_nf.column_dimensions['B'].width = 15
        ws_nf.column_dimensions['C'].width = 40
        ws_nf.column_dimensions['D'].width = 40
    
    # 创建仅条码列表工作表（用于直接处理）
    ws_barcodes = wb.create_sheet(title='条码列表')
    
    # 写入表头
    cell = ws_barcodes.cell(row=1, column=1, value='条码')
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')
    
    # 写入所有条码
    row_num = 2
    for record in na_records + not_found_records:
        ws_barcodes.cell(row=row_num, column=1, value=record['barcode'])
        row_num += 1
    
    ws_barcodes.column_dimensions['A'].width = 15
    
    # 保存文件
    wb.save(filename)
    print(f"✓ Excel 清单已导出到: {filename}")
    print(f"  - 可直接将此文件用于 backend.py 重新处理")

def export_to_json(na_records, not_found_records, filename='old_failures_list.json'):
    """导出清单到 JSON 文件"""
    data = {
        'export_time': datetime.now().isoformat(),
        'total_count': len(na_records) + len(not_found_records),
        'na_count': len(na_records),
        'not_found_count': len(not_found_records),
        'na_records': na_records,
        'not_found_records': not_found_records
    }
    
    with open(filename, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    
    print(f"✓ 清单已导出到: {filename}")

def export_to_txt(na_records, not_found_records, filename='old_failures_list.txt'):
    """导出清单到文本文件"""
    with open(filename, 'w', encoding='utf-8') as f:
        f.write("=" * 80 + "\n")
        f.write("旧失败记录清单\n")
        f.write(f"导出时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write("=" * 80 + "\n\n")
        
        f.write(f"总计: {len(na_records) + len(not_found_records)} 条记录\n\n")
        
        # N/A 记录
        f.write(f"【N/A 记录】({len(na_records)} 条)\n")
        f.write("-" * 80 + "\n")
        for i, record in enumerate(na_records, 1):
            f.write(f"{i}. 条码: {record['barcode']}\n")
            f.write(f"   图片URL: {record['image_url']}\n")
            f.write(f"   图片路径: {record['image_filepath']}\n")
            f.write("\n")
        
        # Not Found 记录
        f.write(f"\n【Not Found 记录】({len(not_found_records)} 条)\n")
        f.write("-" * 80 + "\n")
        for i, record in enumerate(not_found_records, 1):
            f.write(f"{i}. 条码: {record['barcode']}\n")
            f.write(f"   图片URL: {record['image_url']}\n")
            f.write(f"   图片路径: {record['image_filepath']}\n")
            f.write("\n")
    
    print(f"✓ 清单已导出到: {filename}")

def export_barcodes_only(na_records, not_found_records, filename='old_failures_barcodes.txt'):
    """仅导出条码列表（每行一个条码）"""
    with open(filename, 'w', encoding='utf-8') as f:
        f.write("# N/A 记录条码\n")
        for record in na_records:
            f.write(f"{record['barcode']}\n")
        
        f.write("\n# Not Found 记录条码\n")
        for record in not_found_records:
            f.write(f"{record['barcode']}\n")
    
    print(f"✓ 条码列表已导出到: {filename}")

def delete_old_failures():
    """删除旧的失败记录，以便重新检索"""
    print("\n⚠️  警告：此操作将从数据库中删除所有 N/A 和 Not Found 记录")
    print("删除后，backend.py 将重新检索这些条码（包括使用 DDGS 备用搜索）")
    
    confirm = input("\n确认删除？(yes/no): ").strip().lower()
    
    if confirm != 'yes':
        print("操作已取消")
        return False
    
    conn = sqlite3.connect(DATABASE_NAME)
    cursor = conn.cursor()
    
    try:
        # 删除 N/A 和 Not Found 记录
        cursor.execute('''
            DELETE FROM products 
            WHERE product_name IN ('N/A', 'Not Found')
        ''')
        
        deleted_count = cursor.rowcount
        conn.commit()
        
        print(f"\n✓ 成功删除 {deleted_count} 条旧失败记录")
        print("这些条码将在下次运行 backend.py 时重新检索")
        return True
        
    except Exception as e:
        print(f"错误：删除失败：{e}")
        conn.rollback()
        return False
    finally:
        conn.close()

def show_sample_records(na_records, not_found_records, limit=10):
    """显示部分记录样本"""
    print("\n【N/A 记录样本】")
    print("-" * 80)
    for i, record in enumerate(na_records[:limit], 1):
        print(f"{i}. 条码: {record['barcode']}")
        print(f"   图片路径: {record['image_filepath']}")
    
    if len(na_records) > limit:
        print(f"... 还有 {len(na_records) - limit} 条 N/A 记录")
    
    print("\n【Not Found 记录样本】")
    print("-" * 80)
    for i, record in enumerate(not_found_records[:limit], 1):
        print(f"{i}. 条码: {record['barcode']}")
        print(f"   图片路径: {record['image_filepath']}")
    
    if len(not_found_records) > limit:
        print(f"... 还有 {len(not_found_records) - limit} 条 Not Found 记录")

def main():
    """主函数"""
    print("=" * 80)
    print("旧失败记录清单工具")
    print("=" * 80)
    print()
    
    # 获取记录
    print("正在查询数据库...")
    records = get_old_failure_records()
    
    if not records:
        print("✓ 没有发现需要重新检索的旧记录")
        return
    
    # 分类记录
    na_records, not_found_records = categorize_records(records)
    
    # 打印统计
    print_summary(na_records, not_found_records)
    
    # 显示样本
    show_sample_records(na_records, not_found_records)
    
    # 菜单
    while True:
        print("\n请选择操作:")
        print("1. 导出到 Excel 文件（推荐，可直接用于系统处理）")
        print("2. 导出详细清单到 JSON 文件")
        print("3. 导出详细清单到 TXT 文件")
        print("4. 仅导出条码列表到 TXT")
        print("5. 导出所有格式")
        print("6. 删除这些记录（以便重新检索）")
        print("0. 退出")
        
        choice = input("\n请输入选项 (0-6): ").strip()
        
        if choice == '1':
            export_to_excel(na_records, not_found_records)
        elif choice == '2':
            export_to_json(na_records, not_found_records)
        elif choice == '3':
            export_to_txt(na_records, not_found_records)
        elif choice == '4':
            export_barcodes_only(na_records, not_found_records)
        elif choice == '5':
            export_to_excel(na_records, not_found_records)
            export_to_json(na_records, not_found_records)
            export_to_txt(na_records, not_found_records)
            export_barcodes_only(na_records, not_found_records)
            print("\n✓ 所有格式已导出")
        elif choice == '6':
            if delete_old_failures():
                print("\n提示：现在可以运行 backend.py 重新检索这些条码")
                break
        elif choice == '0':
            print("已退出")
            break
        else:
            print("无效选项，请重新选择")

if __name__ == "__main__":
    main()
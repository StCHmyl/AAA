import sqlite3
import openpyxl
from openpyxl.utils import get_column_letter
import os

def export_db_to_excel(db_path, output_file='db_export.xlsx'):
    # 连接SQLite数据库
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    
    # 获取所有表名
    cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
    tables = cursor.fetchall()
    
    # 创建Excel工作簿
    wb = openpyxl.Workbook()
    
    for table in tables:
        table_name = table[0]
        # 获取表数据
        cursor.execute(f"SELECT * FROM {table_name}")
        data = cursor.fetchall()
        
        # 获取列名
        cursor.execute(f"PRAGMA table_info({table_name})")
        columns = [column[1] for column in cursor.fetchall()]
        
        # 创建工作表
        ws = wb.create_sheet(title=table_name)
        
        # 写入列名
        for col_num, column in enumerate(columns, 1):
            ws.cell(row=1, column=col_num, value=column)
        
        # 写入数据
        for row_num, row_data in enumerate(data, 2):
            for col_num, cell_data in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=cell_data)
    
    # 删除默认创建的空工作表
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    # 保存Excel文件
    wb.save(output_file)
    conn.close()
    print(f"数据库已成功导出到 {output_file}")

if __name__ == "__main__":
    db_file = 'barcode_cache.db'
    if not os.path.exists(db_file):
        print(f"错误: 数据库文件 {db_file} 不存在")
    else:
        export_db_to_excel(db_file)
